"""Excel import/export helpers for column-oriented data."""

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

ExcelTable = dict[str, tuple[Any, ...]]


def export_excel(file_name: str | Path, data: ExcelTable) -> None:
    """Export column-oriented data to an Excel workbook."""
    if not data:
        raise ValueError("Excel export data cannot be empty")

    headers = list(data.keys())
    row_count = max((len(values) for values in data.values()), default=0)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(headers)

    for row_index in range(row_count):
        worksheet.append(
            [
                values[row_index] if row_index < len(values) else None
                for values in data.values()
            ]
        )

    workbook.save(file_name)


def import_excel(file_name: str | Path) -> ExcelTable:
    """Import the first worksheet from an Excel workbook as column-oriented data."""
    workbook = load_workbook(file_name, data_only=True, read_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        workbook.close()
        raise ValueError("Excel file is empty") from exc

    headers = [
        str(header).strip() if header is not None else "" for header in header_row
    ]
    if not any(headers):
        workbook.close()
        raise ValueError("Excel file does not contain column headers")

    named_headers = [header for header in headers if header]
    if len(named_headers) != len(set(named_headers)):
        workbook.close()
        raise ValueError("Excel file contains duplicate column headers")

    table: dict[str, list[Any]] = {header: [] for header in named_headers}
    for row in rows:
        if not any(cell is not None and cell != "" for cell in row):
            continue

        for column_index, header in enumerate(headers):
            if not header:
                continue
            value = row[column_index] if column_index < len(row) else None
            table[header].append(value)

    workbook.close()
    return {header: tuple(values) for header, values in table.items()}
